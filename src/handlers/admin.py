from aiogram import Dispatcher, types
from aiogram.contrib.fsm_storage.memory import MemoryStorage
from aiogram.dispatcher import FSMContext
from aiogram.dispatcher import filters
from aiogram.dispatcher.filters.state import State, StatesGroup
from aiogram.utils import executor
from services.database import get_database_session, User, Base, Referral
from keyboards.admin_kb import admin_inlin_kb
from utils.admin_utils import is_admin
import openpyxl
from openpyxl.styles import Font
import os
from datetime import datetime, timedelta
from sqlalchemy import func, desc, text
import asyncio
from services.user_service import UserService  # Import UserService from the appropriate module
from services.channel_service import ChannelService  # Import ChannelService from the appropriate module

# States for admin operations
class AdminStates(StatesGroup):
    waiting_for_search = State()
    waiting_for_block_username = State()
    waiting_for_unblock_username = State()
    waiting_for_mass_message = State()
    browsing_letters = State()  # New: for letter grid view
    browsing_users_by_letter = State()  # New: for viewing users by letter
    waiting_for_channel_input = State()  # New: for adding channel

async def admin_panel(message: types.Message):
    # Проверяем, является ли пользователь администратором
    if is_admin(message.from_user.id):
        await message.answer("Панель администратора:", reply_markup=admin_inlin_kb)
    else:
        await message.answer("У вас нет прав доступа к панели администратора.")

async def admin_callback_handler(callback: types.CallbackQuery, state: FSMContext):
    if not is_admin(callback.from_user.id):
        await callback.answer("У вас нет прав доступа!", show_alert=True)
        return
    
    # Answer callback immediately to prevent timeout
    try:
        await callback.answer()
    except Exception as e:
        print(f"Error answering callback: {e}")
    
    # Save the original message to delete it later
    orig_message = callback.message
    
    # Обработка различных callback_data для админ-панели
    if callback.data == "user_stats":
        await orig_message.delete()  # Delete the previous message
        await view_user_statistics(orig_message)
    elif callback.data == "export_users":
        await orig_message.delete()  # Delete the previous message
        await export_user_list(orig_message)
    elif callback.data == "search_user":
        await orig_message.delete()  # Delete the previous message
        # Create keyboard with search options
        keyboard = types.InlineKeyboardMarkup()
        keyboard.add(
            types.InlineKeyboardButton("🔍 Поиск по тексту", callback_data="text_search"),
            types.InlineKeyboardButton("🔤 Поиск по алфавиту", callback_data="letter_search")
        )
        keyboard.add(types.InlineKeyboardButton("◀️ Назад", callback_data="admin_back"))
        await orig_message.answer("Выберите тип поиска:", reply_markup=keyboard)
    elif callback.data == "block_user":
        await orig_message.delete()  # Delete the previous message
        # Add back button to block prompt
        keyboard = types.InlineKeyboardMarkup()
        keyboard.add(types.InlineKeyboardButton("◀️ Назад", callback_data="cancel_state"))
        await orig_message.answer("Введите имя пользователя (username) для блокировки:", reply_markup=keyboard)
        await AdminStates.waiting_for_block_username.set()
    elif callback.data == "unblock_user":
        await orig_message.delete()  # Delete the previous message
        # Add back button to unblock prompt
        keyboard = types.InlineKeyboardMarkup()
        keyboard.add(types.InlineKeyboardButton("◀️ Назад", callback_data="cancel_state"))
        await orig_message.answer("Введите имя пользователя (username) для разблокировки:", reply_markup=keyboard)
        await AdminStates.waiting_for_unblock_username.set()
    elif callback.data == "mass_message":
        await orig_message.delete()  # Delete the previous message
        # Add back button to mass message prompt
        keyboard = types.InlineKeyboardMarkup()
        keyboard.add(types.InlineKeyboardButton("◀️ Назад", callback_data="cancel_state"))
        await orig_message.answer("Введите сообщение для массовой рассылки:", reply_markup=keyboard)
        await AdminStates.waiting_for_mass_message.set()
    elif callback.data == "manage_channels":
        await orig_message.delete()  # Delete the previous message
        # Remove the keyboard creation here, as it's already created inside the function
        await manage_channels_menu(orig_message)
    elif callback.data == "referral_stats":
        await orig_message.delete()  # Delete the previous message
        await view_referral_statistics(orig_message)
    elif callback.data == "admin_ref_link":
        await orig_message.delete()  # Delete the previous message
        await admin_referral_link(orig_message)
    elif callback.data == "admin_my_refs":
        await orig_message.delete()  # Delete the previous message
        await admin_my_referrals(orig_message)
    elif callback.data == "admin_back":
        # Return to main admin panel with a fresh message
        await orig_message.delete()  # Already deleting the message
        await orig_message.answer("Панель администратора:", reply_markup=admin_inlin_kb)
    elif callback.data == "cancel_state":
        # Cancel any active state
        current_state = await state.get_state()
        if current_state:
            await state.finish()
        await orig_message.delete()  # Already deleting the message
        await orig_message.answer("Действие отменено.", reply_markup=admin_inlin_kb)
    elif callback.data == "text_search":
        await text_search_handler(callback, state)
    elif callback.data == "letter_search":
        await letter_search_handler(callback, state)

async def text_search_handler(callback: types.CallbackQuery, state: FSMContext):
    """Handle text-based user search"""
    if not is_admin(callback.from_user.id):
        await callback.answer("У вас нет прав доступа!", show_alert=True)
        return
    
    try:
        await callback.answer()
    except Exception as e:
        print(f"Error answering callback: {e}")
    
    # Delete the previous message
    await callback.message.delete()
    
    # Show search prompt with back button
    keyboard = types.InlineKeyboardMarkup()
    # Change from cancel_state to search_user to go back one step
    keyboard.add(types.InlineKeyboardButton("◀️ Назад", callback_data="search_user"))
    await callback.message.answer("Введите имя пользователя (username) для поиска или userid:", reply_markup=keyboard)
    
    # Set state for waiting for search input
    await AdminStates.waiting_for_search.set()

async def letter_search_handler(callback: types.CallbackQuery, state: FSMContext):
    """Handle letter-based user search"""
    if not is_admin(callback.from_user.id):
        await callback.answer("У вас нет прав доступа!", show_alert=True)
        return
    
    try:
        await callback.answer()
    except Exception as e:
        print(f"Error answering callback: {e}")
    
    # Clear any existing state first - ADD THIS LINE
    current_state = await state.get_state()
    if current_state is not None:
        await state.finish()
    
    # Delete the previous message
    await callback.message.delete()
    
    # Query database for all users
    session = get_database_session()
    try:
        users = session.query(User).all()
        
        # Extract first letters of usernames (case insensitive)
        first_letters = set()
        for user in users:
            if user.username and len(user.username) > 0:
                first_letter = user.username[0].upper()
                first_letters.add(first_letter)
        
        # Sort letters alphabetically
        first_letters = sorted(list(first_letters))
        
        # Create keyboard with letter buttons (4 columns)
        keyboard = types.InlineKeyboardMarkup(row_width=4)
        letter_buttons = []
        
        for letter in first_letters:
            letter_buttons.append(types.InlineKeyboardButton(
                letter, callback_data=f"letter_{letter}"
            ))
        
        # Add all buttons in groups of 4
        for i in range(0, len(letter_buttons), 4):
            chunk = letter_buttons[i:min(i+4, len(letter_buttons))]
            keyboard.add(*chunk)
        
        # Add back button
        keyboard.add(types.InlineKeyboardButton("◀️ Назад", callback_data="search_user"))
        
        await callback.message.answer("Выберите первую букву имени пользователя:", reply_markup=keyboard)
        
        # Set state for letter browsing
        await AdminStates.browsing_letters.set()
        
    except Exception as e:
        await callback.message.answer(f"Error getting users: {str(e)}")
    finally:
        session.close()

# Fix the letter_select_handler function to ensure it only processes actual letters
async def letter_select_handler(callback: types.CallbackQuery, state: FSMContext):
    """Handle selection of a letter to view users"""
    if not is_admin(callback.from_user.id):
        await callback.answer("У вас нет прав доступа!", show_alert=True)
        return
    
    try:
        await callback.answer()
    except Exception as e:
        print(f"Error answering callback: {e}")
    
    # Get the selected letter
    callback_data = callback.data
    
    # Make sure this is actually a letter callback and not another command
    if not callback_data.startswith("letter_") or len(callback_data) < 8:
        print(f"Invalid letter selection: {callback_data}")
        return
    
    letter = callback_data.replace("letter_", "")
    
    # Additional validation - ensure it's a single character letter
    if len(letter) != 1:
        print(f"Invalid letter: {letter}")
        return
    
    # Store the letter and current page in state
    await state.update_data(letter=letter, page=0)
    
    # Delete the previous message
    await callback.message.delete()
    
    # Show users with selected first letter
    await show_users_by_letter(callback.message, letter, 0, state)

async def show_users_by_letter(message, letter, page, state):
    """Show users with names starting with the specified letter"""
    session = get_database_session()
    try:
        # Query users whose username starts with the letter (case insensitive)
        users_query = session.query(User).filter(
            func.upper(func.substr(User.username, 1, 1)) == letter.upper()
        ).order_by(User.username)
        
        # Get total count for pagination
        total_users = users_query.count()
        
        # Calculate pagination
        users_per_page = 7
        total_pages = max(1, (total_users + users_per_page - 1) // users_per_page)
        
        # Ensure page is in valid range
        page = max(0, min(page, total_pages - 1))
        
        # Update page in state
        await state.update_data(page=page)
        
        # Get users for current page
        users = users_query.offset(page * users_per_page).limit(users_per_page).all()
        
        # Create keyboard with user buttons
        keyboard = types.InlineKeyboardMarkup(row_width=1)
        
        # Add user buttons
        for user in users:
            display_name = f"{user.full_name} (@{user.username})"
            keyboard.add(types.InlineKeyboardButton(
                display_name, callback_data=f"view_user_{user.id}"
            ))
        
        # Add navigation buttons
        nav_buttons = []
        
        # Previous button (if not on first page)
        if page > 0:
            nav_buttons.append(types.InlineKeyboardButton(
                "⬅️ Назад", callback_data="prev_page"
            ))
        
        # Page indicator
        nav_buttons.append(types.InlineKeyboardButton(
            f"{page + 1}/{total_pages}", callback_data="page_info"
        ))
        
        # Next button (if not on last page)
        if page < total_pages - 1:
            nav_buttons.append(types.InlineKeyboardButton(
                "➡️ Далее", callback_data="next_page"
            ))
        
        # Add navigation row
        keyboard.row(*nav_buttons)
        
        # Add back buttons
        keyboard.add(types.InlineKeyboardButton("🔤 К выбору буквы", callback_data="letter_search"))
        keyboard.add(types.InlineKeyboardButton("🏠 В главное меню", callback_data="admin_back"))
        
        # Show message with users for the selected letter
        await message.answer(
            f"Пользователи на букву '{letter}':\n"
            f"Страница {page + 1} из {total_pages}",
            reply_markup=keyboard
        )
        
        # Set state for browsing users by letter
        await AdminStates.browsing_users_by_letter.set()
        
    except Exception as e:
        await message.answer(f"Error getting users: {str(e)}")
    finally:
        session.close()

async def view_user_statistics(message: types.Message):
    session = get_database_session()
    try:
        # Общее количество пользователей
        total_users = session.query(User).count()
        
        # Добавляем проверку на наличие поля is_blocked
        has_blocked_field = hasattr(User, 'is_blocked')
        
        # Статистика по блокировкам (если поле существует)
        blocked_users = 0
        if has_blocked_field:
            blocked_users = session.query(User).filter(User.is_blocked == True).count()
        
        # Новые пользователи за последние 24 часа (если есть поле created_at)
        new_users_24h = 0
        if hasattr(User, 'created_at'):
            new_users_24h = session.query(User).filter(
                User.created_at >= (datetime.now() - timedelta(days=1))
            ).count()
        
        # Формируем статистику
        stats = f"📊 <b>Статистика пользователей:</b>\n\n" \
                f"👥 Всего пользователей: <b>{total_users}</b>\n"
        
        if has_blocked_field:
            stats += f"🚫 Заблокированных пользователей: <b>{blocked_users}</b>\n"
        
        if hasattr(User, 'created_at'):
            stats += f"🆕 Новых пользователей за 24ч: <b>{new_users_24h}</b>\n"
        
        # Add back button
        keyboard = types.InlineKeyboardMarkup()
        keyboard.add(types.InlineKeyboardButton("◀️ Назад", callback_data="admin_back"))
        
        await message.answer(stats, parse_mode="HTML", reply_markup=keyboard)
        
    except Exception as e:
        await message.answer(f"Ошибка при получении статистики: {str(e)}")
    finally:
        session.close()

async def export_user_list(message: types.Message):
    try:
        session = get_database_session()
        users = session.query(User).all()
        
        # Create Excel workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Users"
        
        # Add headers
        headers = ["ID", "Username", "Full Name"]
        if hasattr(User, 'is_blocked'):
            headers.append("Blocked")
        if hasattr(User, 'is_exception'):
            headers.append("Exception")
        if hasattr(User, 'created_at'):
            headers.append("Registration Date")
            
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.font = Font(bold=True)
        
        # Add user data
        for row_num, user in enumerate(users, 2):
            ws.cell(row=row_num, column=1).value = user.id
            ws.cell(row=row_num, column=2).value = user.username
            ws.cell(row=row_num, column=3).value = user.full_name
            
            col = 4
            if hasattr(User, 'is_blocked'):
                ws.cell(row=row_num, column=col).value = "Yes" if user.is_blocked else "No"
                col += 1
            if hasattr(User, 'is_exception'):
                ws.cell(row=row_num, column=col).value = "Yes" if user.is_exception else "No"
                col += 1
            if hasattr(User, 'created_at'):
                ws.cell(row=row_num, column=col).value = user.created_at.strftime("%Y-%m-%d %H:%M:%S") if user.created_at else ""
        
        # Save file
        filename = f"users_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        filepath = os.path.join(os.getcwd(), filename)
        wb.save(filepath)
        
        # Send file to admin with back button
        keyboard = types.InlineKeyboardMarkup()
        keyboard.add(types.InlineKeyboardButton("◀️ Назад", callback_data="admin_back"))
        
        with open(filepath, 'rb') as file:
            await message.answer_document(
                types.InputFile(file, filename=filename),
                caption=f"Exported {len(users)} users",
                reply_markup=keyboard
            )
        
        # Clean up file
        os.remove(filepath)
        session.close()
        
    except Exception as e:
        # Error message with back button
        keyboard = types.InlineKeyboardMarkup()
        keyboard.add(types.InlineKeyboardButton("◀️ Назад", callback_data="admin_back"))
        await message.answer(f"Error exporting users: {str(e)}", reply_markup=keyboard)

async def search_user(message: types.Message, state: FSMContext):
    search_query = message.text.strip()
    session = get_database_session()
    
    try:
        # Existing user search code...
        user = session.query(User).filter(
            (func.lower(User.username) == func.lower(search_query)) |
            (func.lower(User.full_name).like(f"%{search_query.lower()}%")) |
            (User.id == search_query if search_query.isdigit() else False)
        ).first()

        if user:
            # Format user info with additional fields if they exist
            user_info = (
                f"👤 <b>User Information:</b>\n\n"
                f"ID: <code>{user.id}</code>\n"
                f"Username: @{user.username}\n"
                f"Full Name: {user.full_name}\n"
            )
            
            # Add additional fields if they exist
            if hasattr(User, 'is_blocked'):
                status = "🚫 Blocked" if user.is_blocked else "✅ Active"
                user_info += f"Status: {status}\n"
                
            # Add exception status
            if hasattr(User, 'is_exception'):
                exception_status = "⭐️ Exception" if user.is_exception else "👤 Regular user"
                user_info += f"Access: {exception_status}\n"
                
            if hasattr(User, 'created_at'):
                user_info += f"Registered: {user.created_at.strftime('%Y-%m-%d %H:%M:%S')}\n"
            
            # Create keyboard with action buttons
            keyboard = types.InlineKeyboardMarkup(row_width=2)
            
            # Add block/unblock button based on current status
            if hasattr(User, 'is_blocked'):
                action = "Unblock" if user.is_blocked else "Block"
                keyboard.add(types.InlineKeyboardButton(
                    action, callback_data=f"{'unblock' if user.is_blocked else 'block'}_{user.id}"
                ))
            
            # Add exception toggle button
            if hasattr(User, 'is_exception'):
                exception_action = "Remove Exception" if user.is_exception else "Make Exception"
                keyboard.add(types.InlineKeyboardButton(
                    exception_action, callback_data=f"{'remove_exception' if user.is_exception else 'add_exception'}_{user.id}"
                ))
            
            # Add the referrals button
            keyboard.add(types.InlineKeyboardButton(
                "👥 Рефералы", callback_data=f"view_referrals_{user.id}"
            ))
            
            keyboard.add(types.InlineKeyboardButton("◀️ Назад", callback_data="admin_back"))
            
            await message.answer(user_info, parse_mode="HTML", reply_markup=keyboard)
        else:
            # Not found message with back button
            keyboard = types.InlineKeyboardMarkup()
            keyboard.add(types.InlineKeyboardButton("◀️ Назад", callback_data="admin_back"))
            await message.answer("User not found.", reply_markup=keyboard)
        
        await state.finish()
        
    except Exception as e:
        keyboard = types.InlineKeyboardMarkup()
        keyboard.add(types.InlineKeyboardButton("◀️ Назад", callback_data="admin_back"))
        await message.answer(f"Error searching for user: {str(e)}", reply_markup=keyboard)
        await state.finish()
    finally:
        session.close()

async def block_user(message: types.Message, state: FSMContext):
    username = message.text.strip()
    session = get_database_session()
    
    try:
        # Check if User has is_blocked field
        if not hasattr(User, 'is_blocked'):
            await message.answer("Database schema doesn't support user blocking. Please update your User model.")
            await state.finish()
            return
        
        user = session.query(User).filter(func.lower(User.username) == func.lower(username)).first()
        if user:
            user.is_blocked = True
            session.commit()
            
            # Notify the user that they've been blocked
            try:
                from bot import bot
                await bot.send_message(
                    chat_id=user.id,
                    text="⛔️ Ваш аккаунт был заблокирован администратором. Вы можете использовать кнопку 'ℹ️ Помощь' для обращения в поддержку."
                )
            except Exception as e:
                # If we can't message the user (they blocked the bot, etc.)
                print(f"Couldn't notify user {user.id} about being blocked: {e}")
                
            # Add back button
            keyboard = types.InlineKeyboardMarkup()
            keyboard.add(types.InlineKeyboardButton("◀️ Назад", callback_data="admin_back"))
            await message.answer(f"User @{username} has been blocked and notified.", reply_markup=keyboard)
        else:
            # Not found message with back button
            keyboard = types.InlineKeyboardMarkup()
            keyboard.add(types.InlineKeyboardButton("◀️ Назад", callback_data="admin_back"))
            await message.answer("User not found.", reply_markup=keyboard)
        
        await state.finish()
        
    except Exception as e:
        # Error with back button
        keyboard = types.InlineKeyboardMarkup()
        keyboard.add(types.InlineKeyboardButton("◀️ Назад", callback_data="admin_back"))
        await message.answer(f"Error blocking user: {str(e)}", reply_markup=keyboard)
        await state.finish()
    finally:
        session.close()

async def unblock_user(message: types.Message, state: FSMContext):
    # Similar to block_user, but for unblocking
    # Add back button to all response messages
    username = message.text.strip()
    session = get_database_session()
    
    try:
        # Check if User has is_blocked field
        if not hasattr(User, 'is_blocked'):
            await message.answer("Database schema doesn't support user blocking. Please update your User model.")
            await state.finish()
            return
            
        user = session.query(User).filter(func.lower(User.username) == func.lower(username)).first()
        if user:
            user.is_blocked = False
            session.commit()
            # Add back button
            keyboard = types.InlineKeyboardMarkup()
            keyboard.add(types.InlineKeyboardButton("◀️ Назад", callback_data="admin_back"))
            await message.answer(f"User @{username} has been unblocked.", reply_markup=keyboard)
        else:
            keyboard = types.InlineKeyboardMarkup()
            keyboard.add(types.InlineKeyboardButton("◀️ Назад", callback_data="admin_back"))
            await message.answer("User not found.", reply_markup=keyboard)
        
        await state.finish()
        
    except Exception as e:
        keyboard = types.InlineKeyboardMarkup()
        keyboard.add(types.InlineKeyboardButton("◀️ Назад", callback_data="admin_back"))
        await message.answer(f"Error unblocking user: {str(e)}", reply_markup=keyboard)
        await state.finish()
    finally:
        session.close()

async def mass_message(message: types.Message, state: FSMContext):
    # Add back button to the final message status
    mass_msg_text = message.text
    session = get_database_session()
    
    try:
        # Get bot instance
        from bot import bot
        
        # Get all active users
        query = session.query(User)
        if hasattr(User, 'is_blocked'):
            query = query.filter(User.is_blocked == False)
        
        users = query.all()
        
        if not users:
            await message.answer("No users to send messages to.")
            await state.finish()
            return
        
        # Send progress report
        status_msg = await message.answer(f"Sending message to {len(users)} users... 0%")
        
        # Send messages
        success_count = 0
        fail_count = 0
        
        for i, user in enumerate(users):
            try:
                # Update progress every 10 users
                if i % 10 == 0 and i > 0:
                    progress = int((i / len(users)) * 100)
                    await status_msg.edit_text(f"Sending message to {len(users)} users... {progress}%")
                
                # Send message to user
                await bot.send_message(chat_id=user.id, text=mass_msg_text)
                success_count += 1
                
                # Small delay to avoid hitting rate limits
                await asyncio.sleep(0.05)
                
            except Exception:
                fail_count += 1
        
        # Final report with back button
        keyboard = types.InlineKeyboardMarkup()
        keyboard.add(types.InlineKeyboardButton("◀️ Назад", callback_data="admin_back"))
        
        await status_msg.edit_text(
            f"Message sending complete:\n"
            f"✅ Successfully sent: {success_count}\n"
            f"❌ Failed: {fail_count}",
            reply_markup=keyboard
        )
        
        await state.finish()
        
    except Exception as e:
        keyboard = types.InlineKeyboardMarkup()
        keyboard.add(types.InlineKeyboardButton("◀️ Назад", callback_data="admin_back"))
        await message.answer(f"Error sending mass message: {str(e)}", reply_markup=keyboard)
        await state.finish()
    finally:
        session.close()

# New function to show referral statistics
async def view_referral_statistics(message: types.Message):
    session = get_database_session()
    try:
        # Total referrals count
        total_referrals = session.query(Referral).count()
        
        # Top referrers
        top_referrers_query = session.query(
            Referral.referred_by, 
            func.count(Referral.id).label('count')
        ).group_by(Referral.referred_by).order_by(text('count DESC')).limit(5)
        
        top_referrers = top_referrers_query.all()
        
        # Format statistics message
        stats = f"📊 <b>Статистика рефералов:</b>\n\n" \
                f"👥 Всего рефералов: <b>{total_referrals}</b>\n\n" \
                f"🏆 <b>Топ 5 пригласивших:</b>\n"
        
        for i, (referrer_id, count) in enumerate(top_referrers, 1):
            if referrer_id:
                referrer = session.query(User).filter(User.id == referrer_id).first()
                if referrer:
                    stats += f"{i}. {referrer.full_name} (@{referrer.username}): {count} приглашений\n"
        
        # Add back button
        keyboard = types.InlineKeyboardMarkup()
        keyboard.add(types.InlineKeyboardButton("◀️ Назад", callback_data="admin_back"))
        
        await message.answer(stats, parse_mode="HTML", reply_markup=keyboard)
        
    except Exception as e:
        await message.answer(f"Ошибка при получении статистики рефералов: {str(e)}")
    finally:
        session.close()

# New function to show admin referral link
async def admin_referral_link(message: types.Message):
    """Generate a referral link for admin users"""
    user_id = message.chat.id
    
    # Get bot username for proper link generation
    from bot import bot
    bot_info = await bot.get_me()
    bot_username = bot_info.username
    
    # Create a referral link with admin ID
    referral_link = f"https://t.me/{bot_username}?start=ref_{user_id}"
    
    # Get how many users this admin has referred
    user_service = UserService()
    referral_count = user_service.count_user_referrals(user_id)
    user_service.close_session()
    
    # Create buttons for the admin
    keyboard = types.InlineKeyboardMarkup(row_width=1)
    keyboard.add(
        types.InlineKeyboardButton(
            "📤 Поделиться", 
            switch_inline_query=f"Приглашаю тебя в наш бот! {referral_link}"
        )
    )
    keyboard.add(
        types.InlineKeyboardButton("📊 Мои рефералы", callback_data="admin_my_refs")
    )
    keyboard.add(
        types.InlineKeyboardButton("◀️ Назад", callback_data="admin_back")
    )
    
    await message.answer(
        f"🔗 <b>Ваша админская реферальная ссылка:</b>\n\n"
        f"<code>{referral_link}</code>\n\n"
        f"Поделитесь этой ссылкой с друзьями! Вы пригласили: <b>{referral_count}</b> пользователей",
        parse_mode="HTML",
        reply_markup=keyboard
    )

async def admin_my_referrals(message: types.Message):
    """Show the admin's referrals"""
    user_id = message.chat.id
    
    user_service = UserService()
    referrals = user_service.get_user_referrals(user_id)
    
    if not referrals:
        # Add back button
        keyboard = types.InlineKeyboardMarkup()
        keyboard.add(types.InlineKeyboardButton("◀️ Назад", callback_data="admin_ref_link"))
        
        await message.answer("У вас пока нет приглашённых пользователей.", reply_markup=keyboard)
        user_service.close_session()
        return
    
    # Build referral list
    referral_text = "👥 <b>Ваши приглашённые пользователи:</b>\n\n"
    for i, ref in enumerate(referrals, 1):
        user = user_service.get_user_by_id(ref.user_id)
        if user:
            date_str = ref.created_at.strftime("%d.%m.%Y") if hasattr(ref, 'created_at') else "неизвестно"
            referral_text += f"{i}. {user.full_name} (@{user.username}) - {date_str}\n"
    
    user_service.close_session()
    
    # Add back button
    keyboard = types.InlineKeyboardMarkup()
    keyboard.add(types.InlineKeyboardButton("◀️ Назад", callback_data="admin_ref_link"))
    
    await message.answer(referral_text, parse_mode="HTML", reply_markup=keyboard)

async def copy_ref_link_callback(callback: types.CallbackQuery):
    """Handle the copy referral link button for both users and admins"""
    try:
        await callback.answer("Ссылка скопирована в сообщение ниже")
    except Exception as e:
        print(f"Error answering callback: {e}")
    
    # Extract user ID from callback
    user_id = callback.data.replace("copy_ref_", "")
    
    # Generate link
    from bot import bot
    bot_info = await bot.get_me()
    bot_username = bot_info.username
    
    referral_link = f"https://t.me/{bot_username}?start=ref_{user_id}"
    
    # Create keyboard with back button
    keyboard = types.InlineKeyboardMarkup()
    
    # Add appropriate back button based on whether this is admin or user
    if is_admin(callback.from_user.id):
        keyboard.add(types.InlineKeyboardButton("◀️ Назад", callback_data="admin_ref_link"))
    
    # Send as a separate message for easy copying
    await callback.message.answer(
        f"<code>{referral_link}</code>\n\nСкопируйте эту ссылку и отправьте друзьям",
        parse_mode="HTML",
        reply_markup=keyboard
    )

async def view_user_referrals(callback: types.CallbackQuery):
    """View referrals invited by a specific user"""
    if not is_admin(callback.from_user.id):
        await callback.answer("У вас нет прав доступа!", show_alert=True)
        return
    
    try:
        await callback.answer()
    except Exception as e:
        print(f"Error answering callback: {e}")
    
    # Get user ID from callback data
    user_id = int(callback.data.replace("view_referrals_", ""))
    
    # Save the original message to delete it
    orig_message = callback.message
    
    user_service = UserService()
    
    try:
        # Get the user whose referrals we're viewing
        target_user = user_service.get_user_by_id(user_id)
        if not target_user:
            await callback.message.answer("Пользователь не найден")
            return
        
        # Get all referrals by this user
        referrals = user_service.get_user_referrals(user_id)
        
        # Delete the previous message first
        await orig_message.delete()
        
        if not referrals or len(referrals) == 0:
            # No referrals found - show message with back button
            keyboard = types.InlineKeyboardMarkup()
            keyboard.add(types.InlineKeyboardButton("◀️ Назад к пользователю", callback_data=f"back_to_user_{user_id}"))
            await callback.message.answer(
                f"Пользователь {target_user.full_name} (@{target_user.username}) не пригласил ни одного пользователя.",
                reply_markup=keyboard
            )
            return
        
        # Format referral information
        total_referrals = len(referrals)
        referral_text = (
            f"👥 <b>Рефералы пользователя {target_user.full_name} (@{target_user.username}):</b>\n\n"
            f"Всего приглашено: <b>{total_referrals}</b> пользователей\n\n"
        )
        
        # List details of each referral
        for i, ref in enumerate(referrals, 1):
            referred_user = user_service.get_user_by_id(ref.user_id)
            if referred_user:
                date_str = ref.created_at.strftime("%d.%m.%Y %H:%M") if hasattr(ref, 'created_at') else "неизвестно"
                status = "🚫 Blocked" if hasattr(referred_user, 'is_blocked') and referred_user.is_blocked else "✅ Active"
                
                referral_text += (
                    f"{i}. <b>{referred_user.full_name}</b> (@{referred_user.username})\n"
                    f"   ID: <code>{referred_user.id}</code> | {status} | 📅 {date_str}\n"
                )
        
        # Create keyboard with back button
        keyboard = types.InlineKeyboardMarkup()
        keyboard.add(types.InlineKeyboardButton("◀️ Назад к пользователю", callback_data=f"back_to_user_{user_id}"))
        
        # Send message after deleting previous one
        await callback.message.answer(referral_text, parse_mode="HTML", reply_markup=keyboard)
        
    except Exception as e:
        print(f"Error getting user referrals: {e}")
        await callback.message.answer(f"Ошибка при получении списка рефералов: {str(e)}")
    finally:
        user_service.close_session()

async def back_to_user_handler(callback: types.CallbackQuery):
    """Handle back to user button"""
    if not is_admin(callback.from_user.id):
        await callback.answer("У вас нет прав доступа!", show_alert=True)
        return
    
    try:
        await callback.answer()
    except Exception as e:
        print(f"Error answering callback: {e}")
    
    # Extract user ID from callback
    user_id = callback.data.replace("back_to_user_", "")
    
    # Save the original message to delete it
    orig_message = callback.message
    
    # Delete the previous message
    await orig_message.delete()
    
    # Simply re-search for the user to show their info
    session = get_database_session()
    
    try:
        user = session.query(User).filter(User.id == user_id).first()
        
        if user:
            # Format user info (similar to search_user)
            user_info = (
                f"👤 <b>User Information:</b>\n\n"
                f"ID: <code>{user.id}</code>\n"
                f"Username: @{user.username}\n"
                f"Full Name: {user.full_name}\n"
            )
            
            # Add additional fields if they exist
            if hasattr(User, 'is_blocked'):
                status = "🚫 Blocked" if user.is_blocked else "✅ Active"
                user_info += f"Status: {status}\n"
                
            # Add exception status
            if hasattr(User, 'is_exception'):
                exception_status = "⭐️ Exception" if user.is_exception else "👤 Regular user"
                user_info += f"Access: {exception_status}\n"
                
            if hasattr(User, 'created_at'):
                user_info += f"Registered: {user.created_at.strftime('%Y-%m-%d %H:%M:%S')}\n"
            
            # Create keyboard with action buttons
            keyboard = types.InlineKeyboardMarkup(row_width=2)
            
            # Add block/unblock button based on current status
            if hasattr(User, 'is_blocked'):
                action = "Unblock" if user.is_blocked else "Block"
                keyboard.add(types.InlineKeyboardButton(
                    action, callback_data=f"{'unblock' if user.is_blocked else 'block'}_{user.id}"
                ))
            
            # Add exception toggle button
            if hasattr(User, 'is_exception'):
                exception_action = "Remove Exception" if user.is_exception else "Make Exception"
                keyboard.add(types.InlineKeyboardButton(
                    exception_action, callback_data=f"{'remove_exception' if user.is_exception else 'add_exception'}_{user.id}"
                ))
            
            # Add the referrals button
            keyboard.add(types.InlineKeyboardButton(
                "👥 Рефералы", callback_data=f"view_referrals_{user.id}"
            ))
            
            keyboard.add(types.InlineKeyboardButton("◀️ Назад", callback_data="admin_back"))
            
            await callback.message.answer(user_info, parse_mode="HTML", reply_markup=keyboard)
        else:
            await callback.message.answer("User not found.")
    except Exception as e:
        print(f"Error navigating back to user: {e}")
        await callback.message.answer(f"Error: {str(e)}")
    finally:
        session.close()

# Handle direct commands for admin operations
async def admin_block_cmd(message: types.Message):
    if not is_admin(message.from_user.id):
        return
    
    args = message.get_args()
    if not args:
        # Add back button even here
        keyboard = types.InlineKeyboardMarkup()
        keyboard.add(types.InlineKeyboardButton("◀️ Назад в админпанель", callback_data="admin_back"))
        await message.answer("Please provide a username to block.", reply_markup=keyboard)
        return
    
    # The existing code will now use block_user which has back buttons added
    class SimpleMessage:
        def __init__(self, text):
            self.text = text
            self.answer = message.answer  # Pass the original answer method
    
    simple_message = SimpleMessage(args)
    await block_user(simple_message, None)

async def admin_unblock_cmd(message: types.Message):
    if not is_admin(message.from_user.id):
        return
    
    args = message.get_args()
    if not args:
        await message.answer("Please provide a username to unblock.")
        return
    
    # Similar to block command
    class SimpleMessage:
        def __init__(self, text):
            self.text = text
    
    simple_message = SimpleMessage(args)
    await unblock_user(simple_message, None)

async def page_info_handler(callback: types.CallbackQuery, state: FSMContext):
    """Handle page info button (no action, just prevent error)"""
    await callback.answer("Текущая страница")

async def prev_page_handler(callback: types.CallbackQuery, state: FSMContext):
    """Handle previous page button"""
    if not is_admin(callback.from_user.id):
        await callback.answer("У вас нет прав доступа!", show_alert=True)
        return
    
    try:
        await callback.answer()
    except Exception as e:
        print(f"Error answering callback: {e}")
    
    # Get current state data
    data = await state.get_data()
    letter = data.get("letter", "A")
    page = data.get("page", 0)
    
    # Move to previous page
    page = max(0, page - 1)
    
    # Delete the previous message
    await callback.message.delete()
    
    # Show users for the new page
    await show_users_by_letter(callback.message, letter, page, state)

async def next_page_handler(callback: types.CallbackQuery, state: FSMContext):
    """Handle next page button"""
    if not is_admin(callback.from_user.id):
        await callback.answer("У вас нет прав доступа!", show_alert=True)
        return
    
    try:
        await callback.answer()
    except Exception as e:
        print(f"Error answering callback: {e}")
    
    # Get current state data
    data = await state.get_data()
    letter = data.get("letter", "A")
    page = data.get("page", 0)
    
    # Move to next page (validation happens in show_users_by_letter)
    page += 1
    
    # Delete the previous message
    await callback.message.delete()
    
    # Show users for the new page
    await show_users_by_letter(callback.message, letter, page, state)

async def view_user_handler(callback: types.CallbackQuery, state: FSMContext):
    """Handle view user button - show user details"""
    if not is_admin(callback.from_user.id):
        await callback.answer("У вас нет прав доступа!", show_alert=True)
        return
    
    try:
        await callback.answer()
    except Exception as e:
        print(f"Error answering callback: {e}")
    
    # Get user ID
    user_id = callback.data.replace("view_user_", "")
    
    # Delete the previous message
    await callback.message.delete()
    
    # Create a mock message to reuse the existing search_user logic
    class MockMessage:
        def __init__(self, text, answer_method):
            self.text = text
            self.answer = answer_method
    
    mock_message = MockMessage(user_id, callback.message.answer)
    
    # Clear state as we're now viewing a specific user
    await state.finish()
    
    # Use existing search_user to show user details
    await search_user(mock_message, state)

def register_admin_handlers(dp: Dispatcher):
    # Main admin panel command handler
    dp.register_message_handler(admin_panel, commands=["admin"])
    
    # Main menu callback handlers
    dp.register_callback_query_handler(admin_callback_handler, 
                                      lambda c: c.data in ["user_stats", "export_users", 
                                                          "search_user", "block_user", 
                                                          "unblock_user", "mass_message", 
                                                          "manage_channels", "admin_back",
                                                          "cancel_state", "referral_stats", 
                                                          "admin_ref_link", "admin_my_refs"])
    
    # Special handler for cancelling states - works in any state
    dp.register_callback_query_handler(cancel_state_handler, 
                                       lambda c: c.data == "cancel_state", 
                                       state="*")
    
    # Channel management handlers
    dp.register_callback_query_handler(manage_channels_menu, lambda c: c.data == "manage_channels")
    dp.register_callback_query_handler(list_channels, lambda c: c.data == "list_channels")
    dp.register_callback_query_handler(channel_info, lambda c: c.data and c.data.startswith("channel_info_"))
    dp.register_callback_query_handler(add_channel_start, lambda c: c.data == "add_channel")
    dp.register_message_handler(add_channel_process, state=AdminStates.waiting_for_channel_input)
    dp.register_callback_query_handler(toggle_channel, lambda c: c.data and c.data.startswith("toggle_channel_"))
    dp.register_callback_query_handler(delete_channel_confirm, lambda c: c.data and c.data.startswith("delete_channel_"))
    dp.register_callback_query_handler(delete_channel_process, lambda c: c.data and c.data.startswith("confirm_delete_channel_"))
    dp.register_callback_query_handler(cancel_channel_add, lambda c: c.data == "cancel_channel_add", state="*")
    
    # State-dependent message handlers
    dp.register_message_handler(search_user, state=AdminStates.waiting_for_search)
    dp.register_message_handler(block_user, state=AdminStates.waiting_for_block_username)
    dp.register_message_handler(unblock_user, state=AdminStates.waiting_for_unblock_username)
    dp.register_message_handler(mass_message, state=AdminStates.waiting_for_mass_message)
    
    # Direct command handlers
    dp.register_message_handler(admin_block_cmd, commands=["block"], is_admin=True)
    dp.register_message_handler(admin_unblock_cmd, commands=["unblock"], is_admin=True)
    
    # Special handlers that work regardless of state
    dp.register_callback_query_handler(
        lambda c, state: back_to_admin_panel(c, state),
        lambda c: c.data == "admin_back",
        state="*"  # Works in any state
    )
    
    # Search option handlers - register with high priority
    dp.register_callback_query_handler(
        text_search_handler,
        lambda c: c.data == "text_search",
        state="*"  # This handler works in ANY state
    )
    
    # First register letter_search_handler with high priority
    dp.register_callback_query_handler(
        letter_search_handler,
        lambda c: c.data == "letter_search",
        state="*"  # This handler works in ANY state
    )
    
    # Back button handlers in various states
    dp.register_callback_query_handler(
        lambda c, state: admin_callback_handler(c, state),
        lambda c: c.data == "search_user",
        state=AdminStates.waiting_for_search
    )
    
    dp.register_callback_query_handler(
        lambda c, state: admin_callback_handler(c, state),
        lambda c: c.data == "search_user",
        state=AdminStates.browsing_letters
    )
    
    # Handle user actions (block/unblock/exceptions)
    dp.register_callback_query_handler(
        handle_user_action,
        lambda c: c.data and (c.data.startswith("block_") or 
                             c.data.startswith("unblock_") or
                             c.data.startswith("add_exception_") or
                             c.data.startswith("remove_exception_"))
    )
    
    # Handle referral-related callbacks
    dp.register_callback_query_handler(
        copy_ref_link_callback, 
        lambda c: c.data and c.data.startswith("copy_ref_")
    )
    
    dp.register_callback_query_handler(
        view_user_referrals,
        lambda c: c.data and c.data.startswith("view_referrals_")
    )
    
    dp.register_callback_query_handler(
        back_to_user_handler,
        lambda c: c.data and c.data.startswith("back_to_user_")
    )
    
    # Letter selection and user browsing - with improved validation
    dp.register_callback_query_handler(
        letter_select_handler,
        lambda c: c.data and c.data.startswith("letter_") and len(c.data.replace("letter_", "")) == 1,
        state=AdminStates.browsing_letters
    )
    
    # Pagination handlers for browsing users by letter
    dp.register_callback_query_handler(
        prev_page_handler,
        lambda c: c.data == "prev_page",
        state=AdminStates.browsing_users_by_letter
    )
    
    dp.register_callback_query_handler(
        next_page_handler,
        lambda c: c.data == "next_page",
        state=AdminStates.browsing_users_by_letter
    )
    
    dp.register_callback_query_handler(
        page_info_handler,
        lambda c: c.data == "page_info",
        state=AdminStates.browsing_users_by_letter
    )
    
    dp.register_callback_query_handler(
        view_user_handler,
        lambda c: c.data and c.data.startswith("view_user_"),
        state=AdminStates.browsing_users_by_letter
    )

async def handle_user_action(callback: types.CallbackQuery, state: FSMContext):
    if not is_admin(callback.from_user.id):
        await callback.answer("У вас нет прав доступа!", show_alert=True)
        return

    # Answer callback immediately to prevent timeout
    try:
        await callback.answer()
    except Exception as e:
        print(f"Error answering callback: {e}")

    # Extract action and user_id from callback data
    action_parts = callback.data.split("_")
    action = action_parts[0] if len(action_parts) > 0 else ""
    user_id = action_parts[-1] if len(action_parts) > 1 else ""
    
    session = get_database_session()
    
    try:
        user = session.query(User).filter(User.id == int(user_id)).first()
        if not user:
            await callback.message.answer("User not found!")
            return
            
        status_message = ""
        
        # Update user block status
        if action == "block":
            user.is_blocked = True
            status_message = f"🚫 User {user.username} has been blocked"
            # Try to notify the user
            try:
                from bot import bot
                await bot.send_message(
                    chat_id=user.id,
                    text="⛔️ Ваш аккаунт был заблокирован администратором."
                )
                status_message += " and notified."
            except Exception as e:
                status_message += "."
                print(f"Couldn't notify user {user.id} about being blocked: {e}")
        elif action == "unblock":
            user.is_blocked = False
            status_message = f"✅ User {user.username} has been unblocked."
            
        # Handle exception status
        elif action == "add" and "exception" in callback.data:
            user.is_exception = True
            status_message = f"⭐️ User {user.username} is now an exception"
            
        elif action == "remove" and "exception" in callback.data:
            user.is_exception = False
            status_message = f"👤 User {user.username} is no longer an exception"
            
        session.commit()
        
        # Update the message with new user info
        # Format user info with updated status
        user_info = (
            f"👤 <b>User Information:</b>\n\n"
            f"ID: <code>{user.id}</code>\n"
            f"Username: @{user.username}\n"
            f"Full Name: {user.full_name}\n"
        )
        
        # Add additional fields with updated status
        if hasattr(User, 'is_blocked'):
            status = "🚫 Blocked" if user.is_blocked else "✅ Active"
            user_info += f"Status: {status}\n"
            
        # Add exception status
        if hasattr(User, 'is_exception'):
            exception_status = "⭐️ Exception" if user.is_exception else "👤 Regular user"
            user_info += f"Access: {exception_status}\n"
            
        if hasattr(User, 'created_at'):
            user_info += f"Registered: {user.created_at.strftime('%Y-%m-%d %H:%M:%S')}\n"
        
        # Create updated keyboard with correct action buttons
        keyboard = types.InlineKeyboardMarkup(row_width=2)
        
        # Add block/unblock button based on current status
        if hasattr(User, 'is_blocked'):
            new_action = "Unblock" if user.is_blocked else "Block"
            keyboard.add(types.InlineKeyboardButton(
                new_action, callback_data=f"{'unblock' if user.is_blocked else 'block'}_{user.id}"
            ))
        
        # Add exception toggle button with updated text
        if hasattr(User, 'is_exception'):
            exception_action = "Remove Exception" if user.is_exception else "Make Exception"
            keyboard.add(types.InlineKeyboardButton(
                exception_action, callback_data=f"{'remove_exception' if user.is_exception else 'add_exception'}_{user.id}"
            ))
        
        # Add the referrals button
        keyboard.add(types.InlineKeyboardButton(
            "👥 Рефералы", callback_data=f"view_referrals_{user.id}"
        ))
        
        keyboard.add(types.InlineKeyboardButton("◀️ Назад", callback_data="admin_back"))
        
        # Update the message with new info and keyboard
        await callback.message.edit_text(
            user_info, 
            parse_mode="HTML", 
            reply_markup=keyboard
        )
        
        # Show a temporary notification about the action
        await callback.answer(status_message, show_alert=True)
        
    except Exception as e:
        error_msg = f"Error updating user: {str(e)}"
        print(error_msg)
        await callback.message.answer(error_msg)
    finally:
        session.close()
        
async def cancel_state_handler(callback: types.CallbackQuery, state: FSMContext):
    """Special handler to cancel any state and return to admin panel"""
    try:
        await callback.answer("Действие отменено")
    except Exception as e:
        print(f"Error answering callback: {e}")
    
    # Cancel the state
    await state.finish()
    
    # Delete the prompt message
    await callback.message.delete()
    
    # Return to admin panel
    await callback.message.answer("Панель администратора:", reply_markup=admin_inlin_kb)

async def handle_pagination(callback: types.CallbackQuery, state: FSMContext):
    """Handle pagination for browsing users by letter"""
    data = await state.get_data()
    letter = data.get("letter")
    page = data.get("page", 0)
    
    if callback.data == "prev_page":
        page -= 1
    elif callback.data == "next_page":
        page += 1
    
    await show_users_by_letter(callback.message, letter, page, state)

async def back_to_admin_panel(callback: types.CallbackQuery, state: FSMContext):
    """Universal handler for returning to admin panel from anywhere"""
    if not is_admin(callback.from_user.id):
        await callback.answer("У вас нет прав доступа!", show_alert=True)
        return
    
    try:
        await callback.answer()
    except Exception as e:
        print(f"Error answering callback: {e}")
    
    # Clear any active state
    current_state = await state.get_state()
    if current_state is not None:
        await state.finish()
    
    # Delete the current message
    await callback.message.delete()
    
    # Show admin panel
    await callback.message.answer("Панель администратора:", reply_markup=admin_inlin_kb)
    

async def manage_channels_menu(message: types.Message):
    """Main menu for channel management"""
    keyboard = types.InlineKeyboardMarkup(row_width=1)
    keyboard.add(
        types.InlineKeyboardButton("📋 Список каналов", callback_data="list_channels"),
        types.InlineKeyboardButton("➕ Добавить канал", callback_data="add_channel"),
        types.InlineKeyboardButton("◀️ Назад", callback_data="admin_back")
    )
    await message.answer("🔧 Управление каналами:", reply_markup=keyboard)

async def list_channels(callback: types.CallbackQuery):
    """Show list of all channels with control buttons"""
    try:
        await callback.answer()
    except Exception as e:
        print(f"Error answering callback: {e}")
    
    await callback.message.delete()
    
    channel_service = ChannelService()
    channels = channel_service.get_all_channels()
    
    if not channels:
        keyboard = types.InlineKeyboardMarkup()
        keyboard.add(
            types.InlineKeyboardButton("➕ Добавить канал", callback_data="add_channel"),
            types.InlineKeyboardButton("◀️ Назад", callback_data="manage_channels")
        )
        await callback.message.answer("📂 Список каналов пуст.", reply_markup=keyboard)
        channel_service.close_session()
        return
    
    # Format channel list message
    text = "📋 <b>Список каналов:</b>\n\n"
    
    for i, channel in enumerate(channels, 1):
        status = "✅ Включен" if channel.is_enabled else "⭕ Отключен"
        text += f"{i}. <b>{channel.channel_name}</b> (<code>{channel.channel_id}</code>)\n   Статус: {status}\n\n"
    
    # Add controls for each channel
    keyboard = types.InlineKeyboardMarkup(row_width=2)
    
    for channel in channels:
        # Add channel name as a label
        keyboard.add(types.InlineKeyboardButton(
            f"{channel.channel_name}",
            callback_data=f"channel_info_{channel.id}"
        ))
    
    # Add navigation buttons
    keyboard.add(
        types.InlineKeyboardButton("➕ Добавить канал", callback_data="add_channel"),
        types.InlineKeyboardButton("◀️ Назад", callback_data="manage_channels")
    )
    
    await callback.message.answer(text, parse_mode="HTML", reply_markup=keyboard)
    channel_service.close_session()

async def channel_info(callback: types.CallbackQuery):
    """Show detailed info about a channel"""
    try:
        await callback.answer()
    except Exception as e:
        print(f"Error answering callback: {e}")
    
    channel_id = int(callback.data.replace("channel_info_", ""))
    
    channel_service = ChannelService()
    channel = channel_service.get_channel_by_id_db(channel_id)
    
    if not channel:
        await callback.message.answer("Канал не найден.")
        channel_service.close_session()
        return
    
    # Format channel info
    status = "✅ Включен" if channel.is_enabled else "⭕ Отключен"
    added_date = channel.added_at.strftime("%d.%m.%Y %H:%M") if hasattr(channel, "added_at") else "Неизвестно"
    
    text = (
        f"📌 <b>Информация о канале:</b>\n\n"
        f"Название: <b>{channel.channel_name}</b>\n"
        f"ID: <code>{channel.channel_id}</code>\n"
        f"Статус: {status}\n"
        f"Добавлен: {added_date}\n"
    )
    
    # Create keyboard with actions
    keyboard = types.InlineKeyboardMarkup(row_width=1)
    
    toggle_text = "Отключить ❌" if channel.is_enabled else "Включить ✅"
    keyboard.add(
        types.InlineKeyboardButton(toggle_text, callback_data=f"toggle_channel_{channel.id}"),
        types.InlineKeyboardButton("Удалить канал", callback_data=f"delete_channel_{channel.id}"),
        types.InlineKeyboardButton("◀️ Назад к списку", callback_data="list_channels")
    )
    
    await callback.message.edit_text(text, parse_mode="HTML", reply_markup=keyboard)
    channel_service.close_session()

async def add_channel_start(callback: types.CallbackQuery, state: FSMContext):
    """Start process of adding a new channel"""
    try:
        await callback.answer()
    except Exception as e:
        print(f"Error answering callback: {e}")
    
    await callback.message.delete()
    
    keyboard = types.InlineKeyboardMarkup()
    keyboard.add(types.InlineKeyboardButton("◀️ Отмена", callback_data="cancel_channel_add"))
    
    await callback.message.answer(
        "➕ <b>Добавление нового канала</b>\n\n"
        "Пожалуйста, введите ID канала (например: -1001234567890) или имя канала с @ (например: @channel_name).\n\n"
        "<i>❗ Бот должен быть администратором канала.</i>",
        parse_mode="HTML", 
        reply_markup=keyboard
    )
    
    await AdminStates.waiting_for_channel_input.set()

async def add_channel_process(message: types.Message, state: FSMContext):
    """Process adding a new channel after receiving input"""
    channel_input = message.text.strip()
    
    # Create a keyboard for going back regardless of outcome
    keyboard = types.InlineKeyboardMarkup()
    keyboard.add(types.InlineKeyboardButton("◀️ Назад", callback_data="cancel_channel_add"))
    
    # Check if input is valid
    if not channel_input:
        await message.answer("⚠️ Пожалуйста, введите корректный ID канала или @username.", reply_markup=keyboard)
        return
    
    # Extract channel ID or username
    channel_id = None
    channel_username = None
    
    if channel_input.startswith('@'):
        channel_username = channel_input
        # Will resolve to ID later
    elif channel_input.startswith('-100'):
        # Likely a channel ID
        try:
            channel_id = int(channel_input)
        except ValueError:
            await message.answer("⚠️ Некорректный формат ID канала.", reply_markup=keyboard)
            return
    else:
        # Try to interpret as ID or use as username
        try:
            channel_id = int(channel_input)
        except ValueError:
            # Not a numeric ID, assume it's a username without @
            channel_username = f"@{channel_input}" if not channel_input.startswith('@') else channel_input
    
    # Get bot instance
    from bot import bot
    
    try:
        # Try to get channel info and check if bot is admin
        if channel_username:
            try:
                chat = await bot.get_chat(channel_username)
                channel_id = chat.id
                channel_name = chat.title or channel_username
            except Exception as e:
                await message.answer(
                    f"⚠️ Не удалось найти канал {channel_username}.\nОшибка: {str(e)}",
                    reply_markup=keyboard
                )
                return
        else:
            try:
                chat = await bot.get_chat(channel_id)
                channel_name = chat.title or str(channel_id)
            except Exception as e:
                await message.answer(
                    f"⚠️ Не удалось найти канал с ID {channel_id}.\nОшибка: {str(e)}",
                    reply_markup=keyboard
                )
                return
        
        # Check if bot is admin in the channel
        try:
            # Get bot's ID
            bot_info = await bot.get_me()
            bot_id = bot_info.id
            
            # Check if the bot is an admin
            bot_member = await bot.get_chat_member(chat.id, bot_id)
            is_admin = bot_member.status in ['administrator', 'creator']
            
            if not is_admin:
                await message.answer(
                    f"⚠️ Бот не является администратором канала {channel_name}.\n"
                    "Пожалуйста, добавьте бота в администраторы канала и попробуйте снова.",
                    reply_markup=keyboard
                )
                return
        except Exception as e:
            await message.answer(
                f"⚠️ Не удалось проверить права бота в канале.\nОшибка: {str(e)}",
                reply_markup=keyboard
            )
            return
        
        # Bot is admin, proceed to add the channel
        channel_service = ChannelService()
        
        # Check if channel already exists
        existing_channel = channel_service.get_channel_by_id(str(channel_id))
        if existing_channel:
            await message.answer(
                f"⚠️ Канал '{channel_name}' уже добавлен в список.",
                reply_markup=keyboard
            )
            channel_service.close_session()
            return
        
        # Add channel to database
        try:
            new_channel = channel_service.add_channel(channel_name, str(channel_id))
            await message.answer(
                f"✅ Канал <b>{channel_name}</b> успешно добавлен и включен!\n\n"
                f"Теперь пользователям потребуется подписка на этот канал для использования бота.",
                parse_mode="HTML",
                reply_markup=keyboard
            )
        except Exception as e:
            await message.answer(
                f"⚠️ Ошибка при добавлении канала в базу данных: {str(e)}",
                reply_markup=keyboard
            )
        
        channel_service.close_session()
        
    except Exception as e:
        await message.answer(
            f"⚠️ Произошла ошибка: {str(e)}",
            reply_markup=keyboard
        )
    
    # Clear state
    await state.finish()

async def cancel_channel_add(callback: types.CallbackQuery, state: FSMContext):
    """Handle cancel button during channel add process"""
    try:
        await callback.answer()
    except Exception as e:
        print(f"Error answering callback: {e}")
    
    # Clear any active state
    current_state = await state.get_state()
    if current_state is not None:
        await state.finish()
    
    # Delete the current message 
    await callback.message.delete()
    
    # Show channel management menu
    await manage_channels_menu(callback.message)

async def toggle_channel(callback: types.CallbackQuery):
    """Toggle a channel's enabled status"""
    try:
        await callback.answer()
    except Exception as e:
        print(f"Error answering callback: {e}")
    
    channel_db_id = int(callback.data.replace("toggle_channel_", ""))
    
    channel_service = ChannelService()
    channel = channel_service.get_channel_by_id_db(channel_db_id)
    
    if not channel:
        await callback.message.answer("Канал не найден.")
        channel_service.close_session()
        return
    
    # Toggle channel status
    try:
        channel = channel_service.toggle_channel_by_id(channel_db_id)
        status = "включен ✅" if channel.is_enabled else "отключен ⭕"
        
        await callback.answer(f"Канал {status}", show_alert=True)
        
        # Return to list view with updated data
        await list_channels(callback)
    except Exception as e:
        await callback.message.answer(f"Ошибка при изменении статуса канала: {str(e)}")
    
    channel_service.close_session()

async def delete_channel_confirm(callback: types.CallbackQuery):
    """Show confirmation before deleting a channel"""
    try:
        await callback.answer()
    except Exception as e:
        print(f"Error answering callback: {e}")
    
    channel_db_id = callback.data.replace("delete_channel_", "")
    
    # Get channel info for confirmation
    channel_service = ChannelService()
    channel = channel_service.get_channel_by_id_db(int(channel_db_id))
    
    if not channel:
        await callback.message.answer("Канал не найден.")
        channel_service.close_session()
        return
    
    # Create confirmation keyboard
    keyboard = types.InlineKeyboardMarkup(row_width=2)
    keyboard.add(
        types.InlineKeyboardButton("❌ Да, удалить", callback_data=f"confirm_delete_channel_{channel_db_id}"),
        types.InlineKeyboardButton("✅ Нет, отмена", callback_data="list_channels")
    )
    
    await callback.message.edit_text(
        f"⚠️ <b>Подтверждение удаления</b>\n\n"
        f"Вы уверены, что хотите удалить канал <b>{channel.channel_name}</b>?",
        parse_mode="HTML",
        reply_markup=keyboard
    )
    
    channel_service.close_session()

async def delete_channel_process(callback: types.CallbackQuery):
    """Process channel deletion after confirmation"""
    try:
        await callback.answer()
    except Exception as e:
        print(f"Error answering callback: {e}")
    
    channel_db_id = int(callback.data.replace("confirm_delete_channel_", ""))
    
    channel_service = ChannelService()
    
    try:
        channel = channel_service.get_channel_by_id_db(channel_db_id)
        if not channel:
            await callback.message.answer("Канал не найден.")
            channel_service.close_session()
            return
        
        channel_name = channel.channel_name
        success = channel_service.delete_channel_by_id(channel_db_id)
        
        if success:
            await callback.answer(f"Канал {channel_name} удален.", show_alert=True)
        else:
            await callback.answer("Не удалось удалить канал.", show_alert=True)
        
        # Return to the updated channel list
        await list_channels(callback)
    except Exception as e:
        await callback.message.answer(f"Ошибка при удалении канала: {str(e)}")
    finally:
        channel_service.close_session()