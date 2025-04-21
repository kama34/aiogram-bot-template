import os
import openpyxl
from openpyxl.styles import Font
from datetime import datetime, timedelta
from aiogram import types
from services.database import get_database_session, User

async def view_user_statistics(message: types.Message):
    """Показывает статистику пользователей"""
    session = get_database_session()
    try:
        total_users = session.query(User).count()
        
        has_blocked_field = hasattr(User, 'is_blocked')
        
        blocked_users = 0
        if has_blocked_field:
            blocked_users = session.query(User).filter(User.is_blocked == True).count()
        
        new_users_24h = 0
        if hasattr(User, 'created_at'):
            new_users_24h = session.query(User).filter(
                User.created_at >= (datetime.now() - timedelta(days=1))
            ).count()
        
        stats = f"📊 <b>Статистика пользователей:</b>\n\n" \
                f"👥 Всего пользователей: <b>{total_users}</b>\n"
        
        if has_blocked_field:
            stats += f"🚫 Заблокированных пользователей: <b>{blocked_users}</b>\n"
        
        if hasattr(User, 'created_at'):
            stats += f"🆕 Новых пользователей за 24ч: <b>{new_users_24h}</b>\n"
        
        keyboard = types.InlineKeyboardMarkup()
        keyboard.add(types.InlineKeyboardButton("◀️ Назад", callback_data="admin_back"))
        
        await message.answer(stats, parse_mode="HTML", reply_markup=keyboard)
        
    except Exception as e:
        await message.answer(f"Ошибка при получении статистики: {str(e)}")
    finally:
        session.close()

async def export_user_list(message: types.Message):
    """Экспортирует список пользователей в Excel-файл"""
    try:
        session = get_database_session()
        users = session.query(User).all()
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Users"
        
        headers = ["ID", "Username", "Full Name"]
        if hasattr(User, 'is_blocked'):
            headers.append("Blocked")
        if hasattr(User, 'is_exception'):
            headers.append("Exception")
        if hasattr(User, 'created_at'):
            headers.append("Registration Date")
            
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.font = Font(bold=True)
        
        for row_num, user in enumerate(users, 2):
            ws.cell(row=row_num, column=1).value = user.id
            ws.cell(row=row_num, column=2).value = user.username
            ws.cell(row=row_num, column=3).value = user.full_name
            
            col = 4
            if hasattr(User, 'is_blocked'):
                ws.cell(row=row_num, column=col).value = "Yes" if user.is_blocked else "No"
                col += 1
            if hasattr(User, 'is_exception'):
                ws.cell(row=row_num, column=col).value = "Yes" if user.is_exception else "No"
                col += 1
            if hasattr(User, 'created_at'):
                ws.cell(row=row_num, column=col).value = user.created_at.strftime("%Y-%m-%d %H:%M:%S") if user.created_at else ""
        
        filename = f"users_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        filepath = os.path.join(os.getcwd(), filename)
        wb.save(filepath)
        
        keyboard = types.InlineKeyboardMarkup()
        keyboard.add(types.InlineKeyboardButton("◀️ Назад", callback_data="admin_back"))
        
        with open(filepath, 'rb') as file:
            await message.answer_document(
                types.InputFile(file, filename=filename),
                caption=f"Exported {len(users)} users",
                reply_markup=keyboard
            )
        
        os.remove(filepath)
        session.close()
        
    except Exception as e:
        keyboard = types.InlineKeyboardMarkup()
        keyboard.add(types.InlineKeyboardButton("◀️ Назад", callback_data="admin_back"))
        await message.answer(f"Error exporting users: {str(e)}", reply_markup=keyboard)